import re

import numpy as np
import openpyxl
import pandas as pd


def read_kinetic_series(path, top_left_label="Cycle Nr."):
    """Read a kinetic time-series plate reader export (.xlsx) into a tidy DataFrame.

    Parameters
    ----------
    path : str or Path
        Path to the Excel file exported from the plate reader.
    top_left_label : str
        The value in the top-left cell that marks the beginning of the data
        block (default ``"Cycle Nr."``).

    Returns
    -------
    pd.DataFrame
        Columns: ``row``, ``col`` (int), ``well``, ``time_s``, ``value`` (float).
        ``'OVER'`` values are replaced with ``NaN``.
    """
    df1 = openpyxl.load_workbook(path).active
    frame = []
    start = False
    for row in range(1, df1.max_row):
        if df1.cell(row=row, column=1).value == top_left_label:
            start = True
            well_list = [cell.value for cell in df1[row]]
        if start:
            for i, col in enumerate(df1.iter_cols(1, df1.max_column)):
                well = well_list[i]
                plate_row = well[0]
                plate_col = well[1:]
                cell_val = col[row].value
                # Skip titles
                if cell_val in well_list:
                    continue
                # Pick up cycle
                elif bool(re.match("^Cycle", well)):
                    cycle = cell_val
                # Pick up time markers
                elif bool(re.match("^Time", well)):
                    time = cell_val
                # Check that we are still in the plate + time series, it is not empty
                elif (
                    cell_val != ""
                    and cell_val is not None
                    and bool(re.match("^[A-Z][0-9]*$", well))
                    and bool(re.match("^[0-9]*$", cycle))
                ):
                    frame.append(
                        [plate_row, plate_col, f"{plate_row}{plate_col}", time, cell_val]
                    )
    df = pd.DataFrame(data=frame, columns=["row", "col", "well", "time_s", "value"])

    # Change over values to NaN
    df = df.replace("OVER", np.nan)

    # Change dtypes
    df = df.astype({"col": int, "value": float})
    return df


def read_plate(path):
    """Read a single-read plate export (.xlsx) into a tidy DataFrame.

    The data block is expected to start with a row whose first cell is ``'<>'``.

    Parameters
    ----------
    path : str or Path
        Path to the Excel file.

    Returns
    -------
    pd.DataFrame
        Columns: ``row``, ``col`` (int), ``well``, ``value`` (float).
    """
    df1 = openpyxl.load_workbook(path).active
    frame = []
    start = False
    for row in range(1, df1.max_row):
        if df1.cell(row=row, column=1).value == "<>":
            start = True
            plate_col_list = [cell.value for cell in df1[row]]
        if start:
            for i, col in enumerate(df1.iter_cols(1, df1.max_column)):
                plate_col = plate_col_list[i]
                cell_val = col[row].value
                if col[1].column == 1:
                    plate_row = cell_val
                # Check that we are still in the plate, it is not empty
                elif (
                    cell_val != ""
                    and cell_val is not None
                    and bool(re.match("^[A-Z]$", plate_row))
                ):
                    frame.append(
                        [plate_row, plate_col, f"{plate_row}{plate_col}", cell_val]
                    )
    df = pd.DataFrame(data=frame, columns=["row", "col", "well", "value"])

    # Change over values to NaN
    df = df.replace("OVER", np.nan)

    # Change dtypes
    df = df.astype({"col": int, "value": float})
    return df


def read_scan(path, top_left_label="Wavel."):
    """Read a wavelength-scan plate reader export (.xlsx) into a tidy DataFrame.

    Parameters
    ----------
    path : str or Path
        Path to the Excel file.
    top_left_label : str
        The value in the top-left cell that marks the beginning of the data
        block (default ``"Wavel."``).

    Returns
    -------
    pd.DataFrame
        Columns: ``row``, ``col`` (int), ``well``, ``wavelength_nm``, ``value`` (float).
    """
    df1 = openpyxl.load_workbook(path).active
    frame = []
    start = False
    for row in range(1, df1.max_row):
        if df1.cell(row=row, column=1).value == top_left_label:
            start = True
            well_list = [cell.value for cell in df1[row]]
        if start:
            for i, col in enumerate(df1.iter_cols(1, df1.max_column)):
                well = well_list[i]
                plate_row = well[0]
                plate_col = well[1:]
                cell_val = col[row].value
                # Skip titles
                if cell_val in well_list:
                    continue
                # Pick up wavelength
                elif bool(re.match("^Wave", well)):
                    wavelength = cell_val
                # Check that we are still in the plate + wavelength, it is not empty
                elif (
                    cell_val != ""
                    and cell_val is not None
                    and bool(re.match("^[A-Z][0-9]*$", well))
                    and bool(re.match("^[0-9]*$", str(wavelength)))
                ):
                    frame.append(
                        [
                            plate_row,
                            plate_col,
                            f"{plate_row}{plate_col}",
                            wavelength,
                            cell_val,
                        ]
                    )
    df = pd.DataFrame(
        data=frame, columns=["row", "col", "well", "wavelength_nm", "value"]
    )

    # Change over values to NaN
    df = df.replace("OVER", np.nan)

    # Change dtypes
    df = df.astype({"col": int, "value": float})
    return df


def label(df, row_dict=None, col_dict=None, well_dict=None):
    """Add label columns to a plate DataFrame using layout dictionaries.

    Each dictionary maps a new column name to a mapping of
    ``{label_value: [list_of_row_letters | col_numbers | well_strings]}``.
    The dictionary is inverted internally so you can write layouts naturally.

    Parameters
    ----------
    df : pd.DataFrame
        Must contain ``'row'``, ``'col'``, and/or ``'well'`` columns.
    row_dict : dict, optional
        ``{col_name: {value: [row_letters, ...], ...}}``.
    col_dict : dict, optional
        ``{col_name: {value: [col_numbers, ...], ...}}``.
    well_dict : dict, optional
        ``{col_name: {value: [well_strings, ...], ...}}``.

    Returns
    -------
    pd.DataFrame
        The input DataFrame with new label columns added.
    """
    if row_dict is not None:
        for key in row_dict:
            label_dict = row_dict[key]
            inv_dict = {row: entry for entry in label_dict for row in label_dict[entry]}
            df[key] = df["row"].map(inv_dict)
    if col_dict is not None:
        for key in col_dict:
            label_dict = col_dict[key]
            inv_dict = {col: entry for entry in label_dict for col in label_dict[entry]}
            df[key] = df["col"].map(inv_dict)
    if well_dict is not None:
        for key in well_dict:
            label_dict = well_dict[key]
            inv_dict = {
                well: entry for entry in label_dict for well in label_dict[entry]
            }
            df[key] = df["well"].map(inv_dict)
    return df
